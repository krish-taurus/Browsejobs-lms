#!/usr/bin/env python3
"""Build the whitelabel feature matrix workbook.

This is the document that becomes the contract annexure, so it is a workbook
rather than a PDF: a partner filters it, a lawyer pastes it into a schedule,
and a salesperson edits one cell rather than re-rendering a document.

Every module row is a real module in the codebase (see content.MODULES). Tier
inclusion is a commercial decision, so the ``TIERS`` mapping below is the one
place it is expressed — change it here and the whole sheet follows.

Usage::

    python3 scripts/whitelabel/matrix.py
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

import content as C  # noqa: E402

OUT = HERE.parents[3] / "docs" / "partner" / "04_BrowseJobs_Whitelabel_Feature_Matrix.xlsx"

INK = "0A1220"
TRUST = "1B6DF0"
SKY = "E7F1FE"
LINE = "DCE6F5"
MUTED = "5A6B85"
VERIFY = "0BA860"
PAPER = "F6F9FE"

FONT = "Arial"

# Which groups land in which tier. Commercial policy, stated once.
GROUP_TIER = {
    "Learning management": ("Yes", "Yes", "Yes"),
    "Placement engine": ("Partial", "Yes", "Yes"),
    "Employer workspace": ("—", "—", "Yes"),
    "Institute CRM": ("—", "Yes", "Yes"),
    "Platform & governance": ("Partial", "Yes", "Yes"),
}

# Module-level exceptions to the group default, so the sheet is honest about
# what a Starter partner actually gets.
OVERRIDES = {
    "Voice mock lab": ("—", "Yes", "Yes"),
    "AI Tutor": ("—", "Yes", "Yes"),
    "Job feed & Jobs for You": ("—", "Yes", "Yes"),
    "Apply assist": ("—", "Yes", "Yes"),
    "Placement pipeline": ("—", "Yes", "Yes"),
    "Mentoring": ("—", "Yes", "Yes"),
    "Mock interviews": ("Yes", "Yes", "Yes"),
    "CV generator & vault": ("Yes", "Yes", "Yes"),
    "Interview question bank": ("Yes", "Yes", "Yes"),
    "Multi-tenancy": ("Yes", "Yes", "Yes"),
    "Whitelabel manager": ("Yes", "Yes", "Yes"),
    "Feature flags & plans": ("Yes", "Yes", "Yes"),
    "Audit logging": ("—", "Yes", "Yes"),
    "AI gateway & telemetry": ("—", "Yes", "Yes"),
    "Data requests (DPDP)": ("—", "Yes", "Yes"),
    "Vouchers & monetisation": ("—", "Yes", "Yes"),
    "Certificates": ("Yes", "Yes", "Yes"),
}

thin = Side(style="thin", color=LINE)


def style(cell, *, bold=False, size=10, colour="000000", fill=None, wrap=False, align="left"):
    cell.font = Font(name=FONT, bold=bold, size=size, color=colour)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.border = Border(bottom=thin)


def sheet_matrix(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Feature matrix"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "BrowseJobs Platform — whitelabel feature matrix"
    style(ws["A1"], bold=True, size=15, colour=INK)
    ws["A2"] = (
        f"{C.COMPANY['entity']} · {C.COMPANY['edition']}. "
        "Every module listed exists in the product today; roadmap items are on the Roadmap tab."
    )
    style(ws["A2"], size=9, colour=MUTED)
    ws["A3"] = C.CONFIDENTIAL
    style(ws["A3"], size=8, colour=MUTED)

    headers = ["Area", "Module", "What it does", "Starter", "Growth", "Enterprise"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        style(c, bold=True, size=10, colour="FFFFFF", fill=TRUST, wrap=True,
              align="center" if i >= 4 else "left")

    row = 6
    first_data_row = row
    for group in C.MODULES:
        default = GROUP_TIER.get(group["group"], ("—", "Yes", "Yes"))
        for name, desc in group["items"]:
            tiers = OVERRIDES.get(name, default)
            style(ws.cell(row=row, column=1, value=group["group"]), size=9, colour=MUTED, wrap=True)
            style(ws.cell(row=row, column=2, value=name), bold=True, size=10, colour=INK, wrap=True)
            style(ws.cell(row=row, column=3, value=desc), size=9, colour="1B2A44", wrap=True)
            for j, val in enumerate(tiers, start=4):
                c = ws.cell(row=row, column=j, value=val)
                style(c, size=10, align="center",
                      colour=VERIFY if val == "Yes" else (MUTED if val == "—" else "B4740A"),
                      bold=val == "Yes",
                      fill=PAPER if j == 6 else None)
            row += 1
    last_data_row = row - 1

    # Counts per tier, computed by formula so an edited cell updates the totals.
    ws.cell(row=row + 1, column=3, value="Modules included (Yes)").font = Font(name=FONT, bold=True, size=10)
    for j in range(4, 7):
        col = get_column_letter(j)
        c = ws.cell(
            row=row + 1,
            column=j,
            value=f'=COUNTIF({col}{first_data_row}:{col}{last_data_row},"Yes")',
        )
        style(c, bold=True, size=11, colour=TRUST, align="center")

    ws.cell(row=row + 3, column=1, value="Legend").font = Font(name=FONT, bold=True, size=9)
    for k, (label, meaning) in enumerate([
        ("Yes", "Included in the tier at no extra charge."),
        ("Partial", "Included with limits — see the Notes column and your order form."),
        ("—", "Not included in this tier."),
    ]):
        ws.cell(row=row + 4 + k, column=1, value=label).font = Font(name=FONT, bold=True, size=9)
        ws.cell(row=row + 4 + k, column=2, value=meaning).font = Font(name=FONT, size=9, color=MUTED)

    note = ws.cell(row=row + 8, column=1,
                   value="Tier inclusion is commercial policy and is confirmed on the order form. "
                         "Editable cells are columns D–F.")
    note.font = Font(name=FONT, size=9, italic=True, color=MUTED)

    dv = DataValidation(type="list", formula1='"Yes,Partial,—"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"D{first_data_row}:F{last_data_row}")

    for col, width in zip("ABCDEF", (22, 30, 62, 12, 12, 13)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A6"


def sheet_pricing(wb: Workbook) -> None:
    ws = wb.create_sheet("Pricing")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Pricing — list prices"
    style(ws["A1"], bold=True, size=14, colour=INK)
    ws["A2"] = ("List prices, ex-GST, monthly on an annual commitment. Every figure here is solved "
                "in 08_..._Cost_and_Pricing_Model.xlsx — change it there first, or the margin it "
                "implies is fiction. Yellow cells are quote-specific and set per deal.")
    style(ws["A2"], size=9, colour=MUTED)

    headers = ["Tier", "For", "Licence — India (₹/mo)", "Licence — Intl ($/mo)",
               "Included seats", "Included AI allowance", "Onboarding fee (India)",
               "Negotiated rate"]
    for i, h in enumerate(headers, start=1):
        style(ws.cell(row=4, column=i, value=h), bold=True, size=10, colour="FFFFFF",
              fill=TRUST, wrap=True)

    yellow = PatternFill("solid", fgColor="FFFF00")
    onboarding = "₹1,25,000"
    for r, tier in enumerate(C.TIERS, start=5):
        style(ws.cell(row=r, column=1, value=tier["name"]), bold=True, size=11, colour=INK)
        style(ws.cell(row=r, column=2, value=tier["for"]), size=9, colour=MUTED, wrap=True)
        for col, value in (
            (3, tier["price_inr"]),
            (4, tier["price_usd"]),
            (5, tier["seats"]),
            (6, tier["allowance"]),
            (7, onboarding),
        ):
            style(ws.cell(row=r, column=col, value=value), size=10, colour=INK, wrap=True)
        # The only cell still unset: what this particular partner actually pays.
        c = ws.cell(row=r, column=8, value=C.PRICE_TBD)
        style(c, size=10, colour="0000FF", align="center")
        c.fill = yellow

    row = 5 + len(C.TIERS) + 1
    ws.cell(row=row, column=1, value="Metered, on top of the licence").font = Font(
        name=FONT, bold=True, size=10)
    row += 1
    for i, h in enumerate(["Meter", "Basis", "What it covers", "India", "Intl"], start=1):
        style(ws.cell(row=row, column=i, value=h), bold=True, size=9, colour="FFFFFF",
              fill=TRUST, wrap=True)
    for k, (name, basis, why, inr, usd) in enumerate(C.METERED, start=1):
        rw = row + k
        style(ws.cell(row=rw, column=1, value=name), bold=True, size=9.5, colour=INK)
        style(ws.cell(row=rw, column=2, value=basis), size=9, colour=MUTED, wrap=True)
        style(ws.cell(row=rw, column=3, value=why), size=9, colour="1B2A44", wrap=True)
        style(ws.cell(row=rw, column=4, value=inr), size=9.5, colour=INK, align="right")
        style(ws.cell(row=rw, column=5, value=usd), size=9.5, colour=INK, align="right")

    row = row + len(C.METERED) + 2
    ws.cell(row=row, column=1, value="Billing notes").font = Font(name=FONT, bold=True, size=10)
    for k, n in enumerate(C.BILLING_NOTES):
        ws.cell(row=row + 1 + k, column=1, value=f"• {n}").font = Font(
            name=FONT, size=9, color=MUTED)

    for col, width in zip("ABCDEFGH", (16, 40, 22, 20, 20, 34, 20, 18)):
        ws.column_dimensions[col].width = width


def sheet_roadmap(wb: Workbook) -> None:
    ws = wb.create_sheet("Roadmap")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Roadmap — not available today"
    style(ws["A1"], bold=True, size=14, colour=INK)
    ws["A2"] = ("Listed so nothing on the Feature matrix tab can be misread as pending. "
                "Never quote a dated commitment from this tab.")
    style(ws["A2"], size=9, colour=MUTED)
    for i, h in enumerate(["Item", "What it will do"], start=1):
        style(ws.cell(row=4, column=i, value=h), bold=True, size=10, colour="FFFFFF", fill=TRUST)
    for r, (t, b) in enumerate(C.ROADMAP, start=5):
        style(ws.cell(row=r, column=1, value=t), bold=True, size=10, colour=INK, wrap=True)
        style(ws.cell(row=r, column=2, value=b), size=9, colour="1B2A44", wrap=True)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 78


def sheet_integrations(wb: Workbook) -> None:
    ws = wb.create_sheet("Integrations")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Integrations — and whose account each runs on"
    style(ws["A1"], bold=True, size=14, colour=INK)
    ws["A2"] = ("“Your account” is billed to you directly and never marked up by us. "
                "“Metered” is measured per tenant and billed on at your order-form rate.")
    style(ws["A2"], size=9, colour=MUTED)
    for i, h in enumerate(["Service", "Used for", "Account"], start=1):
        style(ws.cell(row=4, column=i, value=h), bold=True, size=10, colour="FFFFFF", fill=TRUST)
    for r, (s, u, a) in enumerate(C.INTEGRATIONS, start=5):
        style(ws.cell(row=r, column=1, value=s), bold=True, size=10, colour=INK)
        style(ws.cell(row=r, column=2, value=u), size=9, colour="1B2A44", wrap=True)
        style(ws.cell(row=r, column=3, value=a), size=9, colour=MUTED, align="center",
              fill=SKY if a == "Metered" else None)
    for col, width in zip("ABC", (26, 58, 20)):
        ws.column_dimensions[col].width = width


def main() -> None:
    wb = Workbook()
    sheet_matrix(wb)
    sheet_pricing(wb)
    sheet_roadmap(wb)
    sheet_integrations(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"  {OUT.name}  ({OUT.stat().st_size / 1024:.0f} KB)")


if __name__ == "__main__":
    main()
