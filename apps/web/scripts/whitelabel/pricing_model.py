#!/usr/bin/env python3
"""Build the whitelabel cost-and-pricing model workbook.

A pricing recommendation without a model behind it is just an opinion, so this
produces a workbook where every number is a live formula over a single
Assumptions sheet. Change the AI usage per student, or the FX rate, and every
tier's margin updates.

The model separates three cost stacks that behave completely differently:

- **Platform** — cents per student per month. Software economics.
- **AI Assist** (tutor, grading, mock analysis) — tens of cents. Still cheap.
- **AI Voice** — dollars. An order of magnitude above the other two combined.

That separation is the whole point. An earlier version of this model bundled a
voice allowance into the licence fee and every tier came out loss-making at any
price the Indian market would pay. Voice is metered for a reason.

Conventions follow the financial-model standard: blue text for hardcoded
inputs, black for formulas, green for cross-sheet links, yellow fill on the
cells a reader is expected to edit.

Sources for each external rate are cited in the Assumptions sheet, next to the
number they justify. Where a rate is a range, the model takes a stated point in
that range and says so.

Usage::

    python3 scripts/whitelabel/pricing_model.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
OUT = HERE.parents[3] / "docs" / "partner" / "08_BrowseJobs_Whitelabel_Cost_and_Pricing_Model.xlsx"

FONT = "Arial"
INK = "0A1220"
TRUST = "1B6DF0"
MUTED = "5A6B85"
LINE = "DCE6F5"
SKY = "E7F1FE"
PAPER = "F6F9FE"
BLUE_TEXT = "0000FF"      # hardcoded input
GREEN_TEXT = "008000"     # link to another sheet
YELLOW = "FFFF00"         # edit me

thin = Side(style="thin", color=LINE)
INR = '₹#,##0;(₹#,##0);-'
INR2 = '₹#,##0.00;(₹#,##0.00);-'
USD = '$#,##0.00;($#,##0.00);-'
USD4 = '$#,##0.0000;($#,##0.0000);-'
PCT = '0.0%'
NUM = '#,##0'


def cell(ws, ref, value, *, bold=False, size=10, colour="000000", fill=None,
         fmt=None, wrap=False, align=None, border=False):
    c = ws[ref] if isinstance(ref, str) else ref
    c.value = value
    c.font = Font(name=FONT, bold=bold, size=size, color=colour)
    c.alignment = Alignment(horizontal=align or "left", vertical="top", wrap_text=wrap)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        c.number_format = fmt
    if border:
        c.border = Border(bottom=thin)
    return c


def header(ws, row, labels, widths=None):
    for i, h in enumerate(labels, start=1):
        cell(ws, ws.cell(row=row, column=i), h, bold=True, size=10,
             colour="FFFFFF", fill=TRUST, wrap=True)
    ws.row_dimensions[row].height = 30
    if widths:
        for col, w in zip(range(1, len(widths) + 1), widths):
            ws.column_dimensions[get_column_letter(col)].width = w


def title(ws, text, sub):
    cell(ws, "A1", text, bold=True, size=15, colour=INK)
    cell(ws, "A2", sub, size=9, colour=MUTED)
    ws.sheet_view.showGridLines = False


def note(ws, row, text, width=4, height=None):
    """A muted explanatory line merged across the table width."""
    c = cell(ws, ws.cell(row=row, column=1), text, size=9, colour=MUTED, wrap=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    if height:
        ws.row_dimensions[row].height = height
    return c


# ----------------------------------------------------------------------------
# 1 - assumptions
# ----------------------------------------------------------------------------

# (label, value, format, source note). Every externally-sourced rate cites where
# it came from, so a stale number is visible rather than invisible.
ASSUMPTIONS = [
    ("§ Currency and tax", None, None, None),
    ("USD → INR", 95.60, "0.00", "Spot rate, 30 Aug 2026. Re-check before quoting."),
    ("GST on Indian invoices", 0.18, PCT, "Pass-through, not cost. All prices below are ex-GST."),

    ("§ AI — LLM token rates (per 1M tokens, USD)", None, None, None),
    ("Workhorse model input", 1.00, USD4, "Claude Haiku 4.5 — tutor, grading, extraction."),
    ("Workhorse model output", 5.00, USD4, "Claude Haiku 4.5."),
    ("Analysis model input", 2.00, USD4, "Claude Sonnet 5 — mock scoring, reports."),
    ("Analysis model output", 10.00, USD4, "Claude Sonnet 5."),

    ("§ AI Assist — usage per AI-active student per month", None, None, None),
    ("Tutor questions", 30, NUM, "Site claims 400+ questions over a course; 30/mo is that pace."),
    ("Tutor input tokens per question", 4000, NUM, "Syllabus context, largely prompt-cached."),
    ("Tutor output tokens per question", 800, NUM, "MEASURE — replace with your ai_events median."),
    ("Graded items per month", 8, NUM, "Assignments plus quizzes."),
    ("Grading input tokens per item", 5000, NUM, "MEASURE — replace with measured."),
    ("Grading output tokens per item", 1000, NUM, "MEASURE — replace with measured."),
    ("Mock interviews analysed", 4, NUM, "Weekly mocks, per the programme."),
    ("Mock analysis input tokens", 15000, NUM, "Transcript plus rubric."),
    ("Mock analysis output tokens", 2000, NUM, "Per-dimension scores and gap report."),

    ("§ AI Voice — the metered line", None, None, None),
    ("Voice platform cost per minute (USD)", 0.11, USD4, "Vapi/Retell realistic all-in: $0.07–0.15/min."),
    ("Telephony cost per minute (USD)", 0.02, USD4, "Outbound India; confirm with your carrier."),
    ("Voice minutes per mock interview", 15, NUM, "One mock = one 15-minute call."),

    ("§ Messaging", None, None, None),
    ("WhatsApp utility messages per student per month", 12, NUM, "Nudges, reminders, results."),
    ("WhatsApp utility rate (INR, ex-GST)", 0.1150, INR2, "Meta India per-message rate, 2026."),
    ("WhatsApp marketing messages per student per month", 2, NUM, "Campaigns."),
    ("WhatsApp marketing rate (INR, ex-GST)", 0.8631, INR2, "Meta India per-message rate, 2026."),

    ("§ Infrastructure — built up from published unit rates", None, None, None),
    ("Video GB streamed per active student per month", 5.0, "0.00", "~10 hrs at 720p (≈0.5 GB/hr)."),
    ("CDN egress rate per GB (USD)", 0.020, USD4, "bunny.net: Asia standard $0.03, volume from $0.005."),
    ("Recording GB retained per active student", 1.8, "0.00", "4 mocks × 15 min × 0.15 GB, 90-day retention."),
    ("Object storage rate per GB-month (USD)", 0.015, USD4, "Cloudflare R2 Standard, zero egress."),
    ("App compute + DB per active student per month (USD)", 0.050, USD4, "MEASURE — amortise tenant 1's bill."),
    ("Platform base cost per tenant per month (USD)", 20.00, USD, "DB share, monitoring, backups, TLS, WAF."),

    ("§ Utilisation", None, None, None),
    ("Monthly-active share of enrolled seats", 0.65, PCT, "An enrolled seat is not a billable-cost seat every month."),

    ("§ Operations", None, None, None),
    ("Support cost — Starter tenant (INR/mo)", 1500, INR, "Email, next business day. Low touch by design."),
    ("Support cost — Growth tenant (INR/mo)", 5000, INR, "Priority queue with an SLA."),
    ("Support cost — Enterprise tenant (INR/mo)", 15000, INR, "Named account manager, partial allocation."),
    ("Onboarding cost per new tenant (INR, one-time)", 60000, INR, "~6 weeks part-time implementation."),
    ("Data migration cost per source system (USD)", 250.00, USD, "ETL build, dry run, cutover."),

    ("§ Margin targets", None, None, None),
    ("Target gross margin — India", 0.68, PCT, "What the Indian anchor price actually supports. See Tier economics."),
    ("Target gross margin — International", 0.82, PCT, "Higher price, identical cost base."),
    ("Licence discount floor — margin", 0.55, PCT, "Below this, founder sign-off required."),
]


def sheet_assumptions(wb):
    ws = wb.active
    ws.title = "Assumptions"
    title(ws, "Cost & pricing model — assumptions",
          "Every yellow cell is an input. Everything on the other sheets is a formula over this one. "
          "Lines marked MEASURE are estimates standing in for data you can read out of ai_events "
          "and your cloud bill — replace them first.")
    header(ws, 4, ["Assumption", "Value", "Source / note"], [52, 14, 66])

    refs, row = {}, 5
    for label, value, fmt, src in ASSUMPTIONS:
        if value is None:
            cell(ws, ws.cell(row=row, column=1), label, bold=True, size=10,
                 colour=TRUST, fill=PAPER)
            for col in (2, 3):
                cell(ws, ws.cell(row=row, column=col), "", fill=PAPER)
        else:
            cell(ws, ws.cell(row=row, column=1), label, size=10, colour=INK, border=True)
            cell(ws, ws.cell(row=row, column=2), value, size=10, colour=BLUE_TEXT,
                 fill=YELLOW, fmt=fmt, align="right", border=True)
            cell(ws, ws.cell(row=row, column=3), src, size=8.5, colour=MUTED,
                 wrap=True, border=True)
            refs[label] = f"Assumptions!$B${row}"
        row += 1

    note(ws, row + 1,
         "Legend: blue on yellow = input you edit · black = formula · green = link to another sheet.",
         width=3)
    return refs


# ----------------------------------------------------------------------------
# 2 - unit economics, split into three stacks
# ----------------------------------------------------------------------------

def sheet_unit(wb, r):
    ws = wb.create_sheet("Unit economics")
    title(ws, "What one student costs us, per month",
          "Three stacks, three orders of magnitude. This page is why AI Voice is metered and the "
          "platform is not.")
    header(ws, 4, ["Cost line", "Formula basis", "USD / active student / month"], [40, 46, 26])

    out = {}
    row = 5

    def block(heading, lines, key):
        nonlocal row
        cell(ws, ws.cell(row=row, column=1), heading, bold=True, size=10,
             colour=TRUST, fill=PAPER)
        for col in (2, 3):
            cell(ws, ws.cell(row=row, column=col), "", fill=PAPER)
        row += 1
        first = row
        for label, basis, formula in lines:
            cell(ws, ws.cell(row=row, column=1), label, size=10, colour=INK, border=True)
            cell(ws, ws.cell(row=row, column=2), basis, size=8.5, colour=MUTED,
                 wrap=True, border=True)
            cell(ws, ws.cell(row=row, column=3), formula, size=10, fmt=USD4,
                 align="right", border=True)
            row += 1
        cell(ws, ws.cell(row=row, column=1), f"{heading} subtotal", bold=True,
             size=10, colour=INK)
        cell(ws, ws.cell(row=row, column=3), f"=SUM(C{first}:C{row - 1})", bold=True,
             size=11, colour=TRUST, fmt=USD4, align="right")
        out[key] = f"'Unit economics'!$C${row}"
        row += 2

    block("Platform — charged in the licence", [
        ("Video delivery",
         "GB streamed × CDN rate",
         f"={r['Video GB streamed per active student per month']}"
         f"*{r['CDN egress rate per GB (USD)']}"),
        ("Recording storage",
         "GB retained × object storage rate",
         f"={r['Recording GB retained per active student']}"
         f"*{r['Object storage rate per GB-month (USD)']}"),
        ("App compute & database",
         "amortised per active student",
         f"={r['App compute + DB per active student per month (USD)']}"),
        ("WhatsApp messaging",
         "(utility + marketing) ÷ FX",
         f"=({r['WhatsApp utility messages per student per month']}"
         f"*{r['WhatsApp utility rate (INR, ex-GST)']}"
         f"+{r['WhatsApp marketing messages per student per month']}"
         f"*{r['WhatsApp marketing rate (INR, ex-GST)']})/{r['USD → INR']}"),
    ], "platform")

    block("AI Assist — sold as a per-student pack", [
        ("AI tutor",
         "questions × (in×rate + out×rate)",
         f"={r['Tutor questions']}*({r['Tutor input tokens per question']}/1000000"
         f"*{r['Workhorse model input']}+{r['Tutor output tokens per question']}/1000000"
         f"*{r['Workhorse model output']})"),
        ("Grading & quizzes",
         "items × (in×rate + out×rate)",
         f"={r['Graded items per month']}*({r['Grading input tokens per item']}/1000000"
         f"*{r['Workhorse model input']}+{r['Grading output tokens per item']}/1000000"
         f"*{r['Workhorse model output']})"),
        ("Mock interview analysis",
         "mocks × (in×rate + out×rate)",
         f"={r['Mock interviews analysed']}*({r['Mock analysis input tokens']}/1000000"
         f"*{r['Analysis model input']}+{r['Mock analysis output tokens']}/1000000"
         f"*{r['Analysis model output']})"),
    ], "assist")

    block("AI Voice — metered by the minute", [
        ("Voice platform, one 15-min mock",
         "minutes × platform rate",
         f"={r['Voice minutes per mock interview']}"
         f"*{r['Voice platform cost per minute (USD)']}"),
        ("Telephony, one 15-min mock",
         "minutes × telephony rate",
         f"={r['Voice minutes per mock interview']}"
         f"*{r['Telephony cost per minute (USD)']}"),
    ], "voice_mock")

    # The headline comparison.
    cell(ws, ws.cell(row=row, column=1), "Reality check — relative size",
         bold=True, size=11, colour=INK)
    row += 1
    for label, formula in [
        ("One month of platform, per student", f"={out['platform']}"),
        ("One month of AI Assist, per student", f"={out['assist']}"),
        ("ONE 15-minute voice mock", f"={out['voice_mock']}"),
        ("Four voice mocks (the weekly programme)", f"={out['voice_mock']}*4"),
    ]:
        cell(ws, ws.cell(row=row, column=1), label, size=10, colour=INK, border=True)
        cell(ws, ws.cell(row=row, column=3), formula, bold=True, size=10,
             colour=TRUST, fmt=USD4, align="right", border=True)
        row += 1

    out["voice_min"] = (f"({r['Voice platform cost per minute (USD)']}"
                        f"+{r['Telephony cost per minute (USD)']})")

    note(ws, row + 1,
         "A student on the full weekly-mock programme costs more in voice minutes than in everything "
         "else combined, several times over. Halve the token counts above and the total barely moves; "
         "halve the voice minutes and the business changes. That is why the licence covers the "
         "platform, AI Assist is a per-student pack, and voice is billed against metered minutes.",
         width=3, height=48)
    return out


# ----------------------------------------------------------------------------
# 3 - tier economics
# ----------------------------------------------------------------------------

# name, enrolled seats, support-cost assumption label, included AI-Assist
# students, included voice minutes, recommended INR/mo, recommended USD/mo
TIERS = [
    ("Starter", 100, "Support cost — Starter tenant (INR/mo)", 25, 150, 24999, 499),
    ("Growth", 500, "Support cost — Growth tenant (INR/mo)", 100, 400, 74999, 1499),
    ("Enterprise", 1500, "Support cost — Enterprise tenant (INR/mo)", 250, 800, 199999, 3499),
]


def sheet_tiers(wb, r, u):
    ws = wb.create_sheet("Tier economics")
    title(ws, "Tier economics — India and International",
          "Cost is modelled bottom-up; price is a round point chosen against the Benchmarks sheet. "
          "Each tier includes a small AI allowance so the AI sells itself — sized to stay a minor "
          "share of the licence.")
    header(ws, 4, [
        "Tier", "Enrolled seats", "Active seats", "Platform cost", "Tenant cost",
        "Included AI cost", "Total cost / mo (USD)",
        "INDIA ₹/mo", "Margin — India", "INTL $/mo", "Margin — Intl",
    ], [13, 12, 11, 13, 12, 14, 16, 13, 13, 12, 12])

    rows = {}
    for i, (name, seats, support_label, ai_students, voice_min, inr, usd) in enumerate(TIERS):
        rw = 5 + i
        rows[name] = rw
        cell(ws, ws.cell(row=rw, column=1), name, bold=True, size=11, colour=INK, border=True)
        cell(ws, ws.cell(row=rw, column=2), seats, size=10, colour=BLUE_TEXT,
             fill=YELLOW, fmt=NUM, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=3),
             f"=B{rw}*{r['Monthly-active share of enrolled seats']}",
             size=10, fmt=NUM, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=4), f"=C{rw}*{u['platform']}",
             size=10, fmt=USD, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=5),
             f"={r['Platform base cost per tenant per month (USD)']}"
             f"+{r[support_label]}/{r['USD → INR']}",
             size=10, fmt=USD, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=6),
             f"={ai_students}*{u['assist']}+{voice_min}*{u['voice_min']}",
             size=10, fmt=USD, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=7), f"=D{rw}+E{rw}+F{rw}", bold=True, size=11,
             colour=TRUST, fmt=USD, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=8), inr, bold=True, size=11, colour=BLUE_TEXT,
             fill=YELLOW, fmt=INR, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=9),
             f"=IFERROR((H{rw}/{r['USD → INR']}-G{rw})/(H{rw}/{r['USD → INR']}),0)",
             bold=True, size=10, fmt=PCT, align="right", colour=TRUST, border=True)
        cell(ws, ws.cell(row=rw, column=10), usd, bold=True, size=11, colour=BLUE_TEXT,
             fill=YELLOW, fmt=USD, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=11), f"=IFERROR((J{rw}-G{rw})/J{rw},0)",
             bold=True, size=10, fmt=PCT, align="right", colour=TRUST, border=True)

    # What each tier includes, in AI terms, spelled out.
    inc = 5 + len(TIERS) + 1
    cell(ws, ws.cell(row=inc, column=1), "Included AI allowance — state this in the order form",
         bold=True, size=11, colour=INK)
    header(ws, inc + 1, ["Tier", "AI Assist students / mo", "Voice minutes / mo",
                         "= mock interviews / mo", "Allowance as % of licence (India)"])
    for i, (name, _s, _sup, ai_students, voice_min, _inr, _usd) in enumerate(TIERS):
        rw = inc + 2 + i
        src = rows[name]
        cell(ws, ws.cell(row=rw, column=1), name, bold=True, size=10, colour=INK, border=True)
        cell(ws, ws.cell(row=rw, column=2), ai_students, size=10, colour=BLUE_TEXT,
             fill=YELLOW, fmt=NUM, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=3), voice_min, size=10, colour=BLUE_TEXT,
             fill=YELLOW, fmt=NUM, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=4),
             f"=B{rw}*0+{voice_min}/{r['Voice minutes per mock interview']}",
             size=10, fmt=NUM, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=5),
             f"=IFERROR(F{src}/(H{src}/{r['USD → INR']}),0)",
             size=10, fmt=PCT, align="right", colour=TRUST, border=True)

    n = inc + 2 + len(TIERS) + 1
    for k, text in enumerate([
        "Margins here are GROSS — cost of service only. They exclude R&D, sales and G&A. A 68% gross "
        "margin is roughly a 30–40% contribution margin once those are loaded.",
        "India is priced below the International equivalent on purpose. The cost base is dollar-denominated "
        "and identical; the Indian anchor price is not. 68% is what this market supports — do not model "
        "82% in India and then discount into a loss.",
        "GST and payment processing are excluded: both are pass-through on collection, not cost of service.",
        "Discount floor: never take a licence below the margin floor on the Assumptions sheet without "
        "founder sign-off. Annual prepay = 2 months free (16.7%) is the standard concession; give the "
        "discount on term, not on price.",
    ]):
        c = cell(ws, ws.cell(row=n + k, column=1), text, size=9, colour=MUTED, wrap=True)
        ws.merge_cells(start_row=n + k, start_column=1, end_row=n + k, end_column=11)
        ws.row_dimensions[n + k].height = 26
    return rows


# ----------------------------------------------------------------------------
# 4 - the metered layer: AI packs, overage, one-time fees
# ----------------------------------------------------------------------------

def sheet_packs(wb, r, u):
    ws = wb.create_sheet("Packs & fees")
    title(ws, "What we sell on top of the licence",
          "The licence is software margin. This page is where AI consumption is recovered — at a "
          "deliberately lower, honest margin, because it is largely someone else's bill.")

    header(ws, 4, ["Meter", "Our cost (USD)", "INDIA list", "Margin — India",
                   "INTL list (USD)", "Margin — Intl"], [42, 15, 15, 14, 15, 14])

    # label, cost formula, INR price, USD price
    METERS = [
        ("AI Assist — per AI-active student / month", f"={u['assist']}", 199, 3.00),
        ("AI Voice — per minute, our keys", f"={u['voice_min']}", 25, 0.40),
        ("AI Voice — per minute, partner's own keys",
         f"={u['voice_min']}*0+0.005", 6, 0.10),
        ("Extra enrolled seat / month",
         f"={u['platform']}*{r['Monthly-active share of enrolled seats']}", 35, 0.60),
    ]
    for i, (label, cost, inr, usd) in enumerate(METERS):
        rw = 5 + i
        cell(ws, ws.cell(row=rw, column=1), label, size=10, colour=INK, border=True)
        cell(ws, ws.cell(row=rw, column=2), cost, size=10, fmt=USD4, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=3), inr, size=10, colour=BLUE_TEXT, fill=YELLOW,
             fmt=INR2, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=4),
             f"=IFERROR((C{rw}/{r['USD → INR']}-B{rw})/(C{rw}/{r['USD → INR']}),0)",
             bold=True, size=10, fmt=PCT, align="right", colour=TRUST, border=True)
        cell(ws, ws.cell(row=rw, column=5), usd, size=10, colour=BLUE_TEXT, fill=YELLOW,
             fmt=USD4, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=6), f"=IFERROR((E{rw}-B{rw})/E{rw},0)",
             bold=True, size=10, fmt=PCT, align="right", colour=TRUST, border=True)

    n = 5 + len(METERS) + 1
    note(ws, n,
         "Voice margin is structurally thinner than software margin — roughly half of the price is a "
         "third party's invoice. Do not fix that by raising the per-minute price until it stops selling: "
         "fix it by moving high-volume partners onto their own Vapi/Twilio keys, where we charge "
         "orchestration only and their voice bill is their own.", width=6, height=42)

    # Voice blocks - a prepaid ladder, so the discount is visible and bounded.
    b = n + 2
    cell(ws, ws.cell(row=b, column=1), "Prepaid voice blocks — India", bold=True,
         size=11, colour=INK)
    header(ws, b + 1, ["Block", "Minutes", "Price (INR)", "Effective ₹/min",
                       "Our cost (USD)", "Margin"])
    BLOCKS = [("Trial", 500, 12499), ("Standard", 1000, 24999),
              ("Volume", 5000, 114999), ("Institutional", 20000, 419999)]
    for i, (name, mins, price) in enumerate(BLOCKS):
        rw = b + 2 + i
        cell(ws, ws.cell(row=rw, column=1), name, bold=True, size=10, colour=INK, border=True)
        cell(ws, ws.cell(row=rw, column=2), mins, size=10, colour=BLUE_TEXT, fill=YELLOW,
             fmt=NUM, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=3), price, size=10, colour=BLUE_TEXT, fill=YELLOW,
             fmt=INR, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=4), f"=IFERROR(C{rw}/B{rw},0)", size=10,
             fmt=INR2, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=5), f"=B{rw}*{u['voice_min']}", size=10,
             fmt=USD, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=6),
             f"=IFERROR((C{rw}/{r['USD → INR']}-E{rw})/(C{rw}/{r['USD → INR']}),0)",
             bold=True, size=10, fmt=PCT, align="right", colour=TRUST, border=True)

    # One-time fees.
    f = b + 2 + len(BLOCKS) + 1
    cell(ws, ws.cell(row=f, column=1), "One-time fees", bold=True, size=11, colour=INK)
    header(ws, f + 1, ["Fee", "Our cost (USD)", "INDIA list", "Margin — India",
                       "INTL list (USD)", "Margin — Intl"])
    FEES = [
        ("Onboarding & implementation",
         f"={r['Onboarding cost per new tenant (INR, one-time)']}/{r['USD → INR']}",
         125000, 2500),
        ("Data migration, per source system",
         f"={r['Data migration cost per source system (USD)']}", 40000, 999),
        ("Custom domain + branded app store listing", "=300", 49999, 999),
    ]
    for i, (label, cost, inr, usd) in enumerate(FEES):
        rw = f + 2 + i
        cell(ws, ws.cell(row=rw, column=1), label, size=10, colour=INK, border=True)
        cell(ws, ws.cell(row=rw, column=2), cost, size=10, fmt=USD, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=3), inr, size=10, colour=BLUE_TEXT, fill=YELLOW,
             fmt=INR, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=4),
             f"=IFERROR((C{rw}/{r['USD → INR']}-B{rw})/(C{rw}/{r['USD → INR']}),0)",
             bold=True, size=10, fmt=PCT, align="right", colour=TRUST, border=True)
        cell(ws, ws.cell(row=rw, column=5), usd, size=10, colour=BLUE_TEXT, fill=YELLOW,
             fmt=USD, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=6), f"=IFERROR((E{rw}-B{rw})/E{rw},0)",
             bold=True, size=10, fmt=PCT, align="right", colour=TRUST, border=True)

    note(ws, f + 2 + len(FEES) + 1,
         "Onboarding is a deal-closer, not a profit centre. It may be discounted to cost to win a "
         "reference partner — but discount the fee, never the licence, because the licence is what "
         "renews.", width=6, height=28)


# ----------------------------------------------------------------------------
# 5 - benchmarks
# ----------------------------------------------------------------------------

BENCHMARKS = [
    ("INDIA — LMS / institute platforms", None, None, None),
    ("Classplus", "₹3,000 – ₹15,000 / mo",
     "Plus setup ₹15k+ and a transaction commission; SMS/WhatsApp credits extra.",
     "edmingle.com, allcoaching.in (2026)"),
    ("Teachmint", "₹2,500 – ₹1,25,000 / mo",
     "Free tier exists; setup ₹25k–₹2L reported at the top end.",
     "igniterapp.com, feealert.in (2026)"),
    ("Learnyst / Edmingle", "Entry ₹8,000 – ₹50,000 / yr",
     "Entry-level annual plans for small institutes.", "edmingle.com (2026)"),

    ("INDIA — education CRM", None, None, None),
    ("LeadSquared", "₹1,250 – ₹4,500 / user / mo",
     "Lite / Pro / Super, billed annually, ex-GST. Education product is quoted.",
     "Published per-user plans (2026)"),
    ("Meritto", "Priced on application volume",
     "Not per-seat — anchors on forms processed.", "softwaresuggest.com (2026)"),
    ("Market norm", "₹2,000 – ₹5,000 / counsellor / mo",
     "Typical per-seat education CRM band.", "vivelead.com (2026)"),

    ("INTERNATIONAL — LMS", None, None, None),
    ("TalentLMS", "$119 / mo (40 users)", "Free tier to 5 users.", "talentlms.com (2026)"),
    ("LearnWorlds", "$249 / mo and up for white-label",
     "White-label gated to the Learning Center plan.", "learnworlds.com (2026)"),
    ("Docebo", "$10,000 – $100,000+ / yr",
     "Quote-only, mid-market to enterprise.", "selecthub.com (2026)"),
    ("Multi-client training platforms", "£4,000 – £25,000 / yr",
     "Training company running several client portals.", "kourses.com (2026)"),
    ("White-label add-on norm", "$150 – $300 / mo",
     "Full white-label tier; branded mobile app often +$150–200/mo.", "kourses.com (2026)"),

    ("THE STACK A BUYER ACTUALLY REPLACES", None, None, None),
    ("Small institute", "≈ ₹22,500 / mo",
     "Classplus at ₹15,000 + LeadSquared Pro × 3 seats at ₹2,500.",
     "Derived from the rows above"),
    ("Growing institute", "≈ ₹51,000 / mo",
     "Classplus at ₹15,000 + LeadSquared Super × 8 seats at ₹4,500.",
     "Derived from the rows above"),
    ("Multi-branch group", "≈ ₹2,15,000 / mo",
     "Teachmint at ₹1,25,000 + LeadSquared Super × 20 seats at ₹4,500.",
     "Derived from the rows above"),
]


def sheet_benchmarks(wb):
    ws = wb.create_sheet("Benchmarks")
    title(ws, "What the market charges",
          "Gathered August 2026. Ranges are as published or as reported by third-party comparison "
          "sites — treat quote-gated vendors as indicative, not exact.")
    header(ws, 4, ["Vendor / segment", "Published price", "Notes", "Source"], [30, 30, 56, 34])

    row = 5
    for name, price, notes, source in BENCHMARKS:
        if price is None:
            cell(ws, ws.cell(row=row, column=1), name, bold=True, size=10,
                 colour=TRUST, fill=PAPER)
            for col in (2, 3, 4):
                cell(ws, ws.cell(row=row, column=col), "", fill=PAPER)
        else:
            cell(ws, ws.cell(row=row, column=1), name, bold=True, size=10, colour=INK, border=True)
            cell(ws, ws.cell(row=row, column=2), price, size=10, colour=INK, border=True)
            cell(ws, ws.cell(row=row, column=3), notes, size=8.5, colour=MUTED,
                 wrap=True, border=True)
            cell(ws, ws.cell(row=row, column=4), source, size=8, colour=MUTED,
                 wrap=True, border=True)
        row += 1

    note(ws, row + 1,
         "Read across, not down: no vendor here sells LMS + institute CRM + placement + AI mocks as one "
         "product. The last block is the comparison a buyer actually makes — and it is the number our "
         "licence should be quoted against, not the Classplus line on its own.",
         height=42)


# ----------------------------------------------------------------------------
# 6 - sensitivity
# ----------------------------------------------------------------------------

def sheet_sensitivity(wb, r, u, tier_rows):
    ws = wb.create_sheet("Sensitivity")
    title(ws, "What breaks the margin",
          "Growth tier at the recommended India price. Two levers matter, and only two.")

    g = tier_rows["Growth"]
    price_usd = f"('Tier economics'!$H${g}/{r['USD → INR']})"
    platform = f"'Tier economics'!$D${g}"
    tenant = f"'Tier economics'!$E${g}"

    # A - voice given away inside the licence
    cell(ws, "A4", "Lever 1 — voice minutes given away inside the licence",
         bold=True, size=11, colour=INK)
    header(ws, 5, ["Included voice minutes / mo", "Included AI cost (USD)",
                   "Total tenant cost (USD)", "Margin at ₹74,999"], [30, 26, 26, 24])
    assist_included = f"100*{u['assist']}"
    for i, mins in enumerate([0, 400, 1000, 2500, 5000, 10000]):
        rw = 6 + i
        cell(ws, ws.cell(row=rw, column=1), mins, size=10, colour=BLUE_TEXT, fill=YELLOW,
             fmt=NUM, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=2), f"={assist_included}+A{rw}*{u['voice_min']}",
             size=10, fmt=USD, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=3), f"={platform}+{tenant}+B{rw}",
             size=10, fmt=USD, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=4),
             f"=IFERROR(({price_usd}-C{rw})/{price_usd},0)",
             bold=True, size=10, fmt=PCT, align="right", colour=TRUST, border=True)

    n = 13
    note(ws, n,
         "400 minutes is the recommended Growth allowance — about 27 mock interviews a month, enough "
         "for the partner to see the product work. Give away 2,500 and the tier is barely profitable; "
         "give away 5,000 and it is not. This is the single most expensive concession a salesperson can "
         "make, so it belongs in the order form, not in a conversation.", height=42)

    # B - AI-active share of seats
    b = n + 2
    cell(ws, ws.cell(row=b, column=1),
         "Lever 2 — how many enrolled seats are actually active", bold=True, size=11, colour=INK)
    header(ws, b + 1, ["Monthly-active share", "Active seats", "Platform cost (USD)",
                       "Total tenant cost (USD)", "Margin at ₹74,999"])
    seats = f"'Tier economics'!$B${g}"
    included = f"'Tier economics'!$F${g}"
    for i, share in enumerate([0.40, 0.55, 0.65, 0.80, 1.00]):
        rw = b + 2 + i
        cell(ws, ws.cell(row=rw, column=1), share, size=10, colour=BLUE_TEXT, fill=YELLOW,
             fmt=PCT, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=2), f"={seats}*A{rw}", size=10, fmt=NUM,
             align="right", border=True)
        cell(ws, ws.cell(row=rw, column=3), f"=B{rw}*{u['platform']}", size=10,
             fmt=USD, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=4), f"=C{rw}+{tenant}+{included}", size=10,
             fmt=USD, align="right", border=True)
        cell(ws, ws.cell(row=rw, column=5),
             f"=IFERROR(({price_usd}-D{rw})/{price_usd},0)",
             bold=True, size=10, fmt=PCT, align="right", colour=TRUST, border=True)

    note(ws, b + 2 + 5 + 1,
         "Platform cost barely moves with utilisation — even at 100% active the margin holds. That is "
         "the shape of a software business, and it is the reason seats can be sold generously while "
         "minutes cannot.", width=5, height=28)


def main() -> None:
    wb = Workbook()
    r = sheet_assumptions(wb)
    u = sheet_unit(wb, r)
    tier_rows = sheet_tiers(wb, r, u)
    sheet_packs(wb, r, u)
    sheet_benchmarks(wb)
    sheet_sensitivity(wb, r, u, tier_rows)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"  {OUT.name}  ({OUT.stat().st_size / 1024:.0f} KB)")


if __name__ == "__main__":
    main()
